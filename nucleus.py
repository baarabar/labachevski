class Nucleus:
    # Эта функция возводит в степень q все элементы одномерного массива и складывает их
    @staticmethod
    def sump(massive, q):
        c = []
        for i in massive:
            c.append(i ** q)
        return c

    # openab - функция для открытия нужных строк(от строки под номером a, включительно, до b, не включительно)
    @staticmethod
    def openab(a, b):
        from meta import file_name

        c = 0
        dmstr = []
        file = open(file_name)
        for i in file:
            c += 1
            if (c >= a) and (c < b):
                i = i[:-1]
                if i[:2] == 'dm':
                    dmstr.append(i.split(' ')[1:])
                else:
                    dmstr.append(i.split(' '))
            for j in range(0, len(dmstr)):
                for g in range(0, len(dmstr[j])):
                    dmstr[j][g] = float(dmstr[j][g])
        return dmstr

    # Эта функция склеивает несколько массивов
    # Очень полезна при создании таблиц
    @staticmethod
    def connect(massive):
        c = 0
        ll = len(massive[0])
        for i in massive:
            if len(i) == ll:
                pass
            else:
                c = 1
                print('Столбцы не одинаковой длинны(')
        if c == 0:
            massive_end = []
            for i in range(0, len(massive[0])):
                ms = []
                for g in massive:
                    ms.append(g[i])
                massive_end.append(ms)
            return massive_end

    @staticmethod
    def opendm(str_number, name):
        from meta import d_p
        from openpyxl import Workbook

        print('*ВЫЧИСЛЕНИЕ ПОГРЕШНОСТИ ДЛЯ СЕРИИ ИЗМЕРЕНИЙ С ПОМОЩЬЮ ФУНКЦИИ OPENDM*')

        data = Nucleus.openab(str_number, str_number + 1)[0]
        tab = []
        srednee = sum(data) / len(data)
        s = 0
        с = 0
        for g in data:
            с += 1
            s += (g - srednee) ** 2
            d = [с, g, g - srednee, (g - srednee) ** 2]
            tab.append(d)
        sco = (s / (len(data) * (len(data) - 1))) ** 0.5 * Nucleus.coef(len(data), d_p)

        print('')
        print('Среднее значение: ' + str(srednee))
        print('Сумма квадратов разности:' + str(s))
        print('Кол-во измерений: ' + str(len(data)))
        print('Коэфициент Стьюдента: ' + str(Nucleus.coef(len(data), d_p)))
        print('Среднеквадратическая погрешность:' + str(sco))
        print('')
        print('')
        print('')

        wb = Workbook()
        ws = wb.active
        for i in tab:
            ws.append(i)
        wb.save('data/laba_10/output/' + str(name) + '.xlsx')

        end = {
            'n': len(data),
            'sr': srednee,
            's': s,
            'coef': Nucleus.coef(len(data), d_p),
            'pogr': sco,
        }

        return end

    # Находит нужный коэфициент Стьюдента
    @staticmethod
    def coef(self, p):
        coef = [
            [],
            [],
            [6.3137515148, 12.7062047364, 63.6567411629, 636.619249432],
            [2.91998558036, 4.30265272991, 9.92484320092, 31.599054577],
            [2.3533634348, 3.18244630528, 5.84090929976, 12.9239786366],
            [2.13184678134, 2.7764451052, 4.60409487142, 8.61030158138],
            [2.01504837267, 2.57058183661, 4.03214298356, 6.86882663987],
            [1.94318028039, 2.44691184879, 3.70742802132, 5.95881617993],
            [1.89457860506, 2.36462425101, 3.49948329735, 5.40788252098],
            [1.85954803752, 2.30600413503, 3.35538733133, 5.04130543339],
            [1.83311293265, 2.26215716274, 3.24983554402, 4.78091258593],
            [1.81246112281, 2.22813885196, 3.16927266718, 4.5868938587],
            [1.7958848187, 2.20098516008, 3.10580651322, 4.43697933823],
            [1.78228755565, 2.17881282966, 3.05453958834, 4.31779128361],
            [1.77093339599, 2.16036865646, 3.01227583821, 4.22083172771],
            [1.76131013577, 2.14478668792, 2.97684273411, 4.14045411274],
            [1.75305035569, 2.13144954556, 2.94671288334, 4.0727651959],
            [1.74588367628, 2.11990529922, 2.92078162235, 4.0149963326],
            [1.73960672608, 2.10981557783, 2.89823051963, 3.96512626361],
            [1.73406360662, 2.10092204024, 2.87844047271, 3.92164582001],
            [1.72913281152, 2.09302405441, 2.86093460645, 3.88340584948],
            [1.72471824292, 2.08596344727, 2.84533970978, 3.84951627298],
            [1.72074290281, 2.07961384473, 2.83135955802, 3.81927716303],
            [1.71714437438, 2.0738730679, 2.8187560606, 3.79213067089],
            [1.71387152775, 2.06865761042, 2.80733568377, 3.76762680377],
            [1.71088207991, 2.06389856163, 2.79693950477, 3.74539861893],
            [1.70814076125, 2.05953855275, 2.78743581368, 3.72514394948],
            [1.70561791976, 2.05552943864, 2.77871453333, 3.70661174331],
            [1.70328844572, 2.05183051648, 2.77068295712, 3.68959171334],
            [1.70113093427, 2.0484071418, 2.76326245546, 3.67390640062],
            [1.69912702653, 2.04522964213, 2.75638590367, 3.6594050194],
            [1.69726089436, 2.0422724563, 2.74999565357, 3.645958635],
        ]
        if p == 0.9:
            return coef[self][0]
        if p == 0.95:
            return coef[self][1]
        if p == 0.99:
            return coef[self][2]

    # Метод наименьших квадратов
    @staticmethod
    def method_min_sqrt(x_str_number=0, y_str_number=1):
        from meta import file_name
        from meta import d_p
        # Конвертация входных данных из файла в массив из цифр
        file = list(open(file_name))
        x = file[x_str_number][:-1].split(' ')
        for i in range(0, len(x)):
            x[i] = float(x[i])
        y = file[y_str_number][:-1].split(' ')
        for i in range(0, len(y)):
            y[i] = float(y[i])
        if len(x) == len(y):
            n = len(x)
            # Вычисление средних значений
            x_sr = sum(x) / len(x)
            y_sr = sum(y) / len(y)
            x2_sr = sum(Nucleus.sump(x, 2)) / len(x)
            y2_sr = sum(Nucleus.sump(y, 2)) / len(y)

            xy = [x, y]
            xy = Nucleus.connect(xy)
            for i in range(0, len(xy)):
                c = 1
                for g in xy[i]:
                    c *= g
                xy[i] = c
            xy_sr = sum(xy) / len(xy)

            # Вычисление параметров и их погрешностей
            sco_x = x2_sr - x_sr ** 2
            sco_y = y2_sr - y_sr ** 2

            a = (xy_sr - x_sr * y_sr) / sco_x
            b = y_sr - a * x_sr

            sco_a = ((sco_y / sco_x - a ** 2) / n - 2) ** 0.5
            sco_b = sco_a * x2_sr ** 0.5

            # Вывод значений на экран
            print('НАЧАЛО ВЫПОЛНЕНИЯ ФУНЦИИ: "МЕТОД НАИМЕНЬШИХ КВАДРАТОВ"')
            print('')
            print('Среднее значение X и Y соответственно: ' + str(x_sr) + ',' + str(y_sr))
            print('Среднее значение квадратов X и Y соответственно: ' + str(x2_sr) + ',' + str(y2_sr))
            print('Среднее значение произведений XY: ' + str(xy_sr))
            print('Квадрат среднеквадратичного отклонения X: ' + str(sco_x))
            print('Квадрат среднеквадратичного отклонения Y: ' + str(sco_y))
            print('Параметры a и b соответственно: ' + str(a) + ',' + str(b))
            print('Среднеквадратичное отклонение a и b соответственно: ' + str(sco_a) + ',' + str(sco_b))
            print('Погрешность a и b соответственно: ' + str(sco_a * Nucleus.coef(len(x), d_p)) + ',' + str(sco_b * Nucleus.coef(len(x), d_p)))
            print('Кол-во измерений: ' + str(len(x)))
            print('Коэфициент Стьюдента: ' + str(Nucleus.coef(len(x), d_p)))
            print('')
            print('')
            print('')

            # Вывод значений в список
            end = {
                'n': n,
                'x_sr': x_sr,
                'y_sr': y_sr,
                'x2_sr': x2_sr,
                'y2_sr': y2_sr,
                'xy_sr': xy_sr,
                'sco_x': sco_x,
                'sco_y': sco_y,
                'a': a,
                'b': b,
                'sco_a': sco_a,
                'sco_b': sco_b,
                'pogr_a': sco_a * Nucleus.coef(len(x), d_p),
                'pogr_b': sco_b * Nucleus.coef(len(x), d_p),
                'coef': Nucleus.coef(len(x), d_p)
            }
            return end
        else:
            print('У строк неодинаковая длинна')
            return 0

    # Метод парных точек
    @staticmethod
    def paired_point_method(massive, name):
        from openpyxl import Workbook
        from meta import laba_number

        # Конвертация входных данных из файла в массив из цифр
        x = massive[0]
        for i in range(0, len(x)):
            x[i] = float(x[i])
        y = massive[1]
        for i in range(0, len(y)):
            y[i] = float(y[i])

        # Выполнение метода парных точек
        if len(x) == len(y):
            n = len(x)
            data = []
            for g in range(0, int(n/2)):
                stroka = [g+1, int(n/2)+g+1, x[int(n/2)+g]-x[g], y[int(n/2)+g]-y[g], (y[int(n/2)+g]-y[g])/(x[int(n/2)+g]-x[g]), 1, 1]
                data.append(stroka)

            k_sr = 0
            sco = 0
            for g in data:
                k_sr += g[4]
            k_sr /= len(data)
            for g in range(0, len(data)):
                data[g][5] = data[g][4] - k_sr
            for g in range(0, len(data)):
                data[g][6] = data[g][5]**2
                sco += data[g][6]
            s = sco
            sco = (sco/(len(data)*(len(data)-1)))**0.5
            sco_1 = sco
            sco *= Nucleus.coef(len(data), 0.95)

            wb = Workbook()
            ws = wb.active
            for j in data:
                ws.append(j)
            wb.save('data/laba_' + str(laba_number) +'/output/' + str(name) + '.xlsx')

            print('НАЧАЛО ВЫПОЛНЕНИЯ ФУНЦИИ: "МЕТОД ПАРНЫХ ТОЧЕК"')
            print('')
            print('Среднее значение коэффициента: ' + str(k_sr))
            print('Сумма разностей квадратов: ' + str(s))
            print('Среднеквадратичное отклонение: ' + str(sco_1))
            print('Погрешность: ' + str(sco))
            print('Коэфициент Стьюдента: ' + str(Nucleus.coef(len(data), 0.95)))
            print()
            print('Создана таблица с названием - ' + 'data/laba_' + str(laba_number) +'/output/' + str(name) + '.xlsx')
            print()
            print()
            print()

            # Вывод значений в список
            end = {
                'n': len(data),
                'sr': k_sr,
                's': s,
                'sco': sco_1,
                'pogr': sco,
                'coef': Nucleus.coef(len(data), 0.95),
            }

            return end
        else:
            print('иди нахуй')
            return 0

    @staticmethod
    def dm(massive, name):
        from meta import d_p
        from openpyxl import Workbook
        from meta import laba_number

        print('*ВЫЧИСЛЕНИЕ ПОГРЕШНОСТИ ДЛЯ СЕРИИ ИЗМЕРЕНИЙ С ПОМОЩЬЮ ФУНКЦИИ DM*')

        data = massive
        tab = []
        srednee = sum(data) / len(data)
        s = 0
        с = 0
        for g in data:
            с += 1
            s += (g - srednee) ** 2
            d = [с, g, g - srednee, (g - srednee) ** 2]
            tab.append(d)
        sco = (s / (len(data) * (len(data) - 1))) ** 0.5 * Nucleus.coef(len(data), d_p)

        print('')
        print('Среднее значение: ' + str(srednee))
        print('Сумма квадратов разности:' + str(s))
        print('Кол-во измерений: ' + str(len(data)))
        print('Коэфициент Стьюдента: ' + str(Nucleus.coef(len(data), d_p)))
        print('Среднеквадратическая погрешность:' + str(sco))
        print('')
        print('')
        print('')

        wb = Workbook()
        ws = wb.active
        for i in tab:
            ws.append(i)
        wb.save('data/laba_' + str(laba_number) +'/output/' + str(name) + '.xlsx')

        end = {
            'n': len(data),
            'sr': srednee,
            's': s,
            'coef': Nucleus.coef(len(data), d_p),
            'pogr': sco,
        }

        return end
